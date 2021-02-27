def integrate_excel(from_file1_path, to_file2_path, columns_of_interest_from_file1=[], key_from_file1=0, key_to_file2=0,
                    starting_column_file2=0, new_column_name='', add_annotation = "No", add_report = False):
    """
    :param from_file1_path: the path of file that contain information we need to add to the other file
    :param to_file2_path: the path of the file that we will paste information to
    :param columns_of_interest_from_file1: a list of columns numbers to be added from file 1
    :param key_from_file1: the number of the key column in file1 (from_file)
    :param key_to_file2: the number of the key column in file2 (to_file)
    :param starting_column_file2: the number of the column we will start pasting new information into
    :param new_column_name: the name of the new column
    :param add_annotation: if "Gene location" is selected, that would add an extra column after merging the 2 files in a new column

    :return:
    """

    """
    Check User errors
    """
    if len(columns_of_interest_from_file1) < 1:
        return 'You should specify the numbers of columns of interest from file 1 to be added to file 2.'
    for i in columns_of_interest_from_file1:
        if type(i) != int:
            return 'The number of each column of interest should be an integer'
        if i < 0:
            return 'The number of each column of interest should be greater than 0'
    if type(key_to_file2) != int or type(key_from_file1) != int:
        return 'The key column should be the number of the column to be a key to match the 2 files and should be greated than zero'
    if key_to_file2 < 1 or key_from_file1 < 1:
        return 'The number of the key column must be at least 1'



    # load the package that will open the excel files
    import openpyxl

    # open the first file from which we will copy information
    wb1 = openpyxl.load_workbook(from_file1_path, data_only=True)
    sheet1 = wb1.active
    c1 = sheet1.cell

    # open the second file to which we will paste information
    wb2 = openpyxl.load_workbook(to_file2_path, data_only=True)
    sheet2 = wb2.active
    c2 = sheet2.cell

    """
    store the data from file 1 in a dictionary (key is "from_file1_path" key column and value is a list from the columns of interest)
    """
    file1_data = {}
    # get all the data from all rows (i)
    for i in range(2, 1000000):
        key = c1(row=i, column=key_from_file1).value
        if key == None:  # if the ~row is empty (~end of file)
            break
        else:
            file1_data[key] = []
            # Iterate over the columns of interest and append their data in a list called 'value'
            for j in columns_of_interest_from_file1:
                file1_data[key].append(c1(row=i, column=j).value)
    # Now, we have a dictionary "file1_data" to be used to add information to the other file

    """
    Paste the data to file 2
    """

    # First, name the new columns with the name the user entered as an argument
    col = starting_column_file2
    for j in range(0, len(columns_of_interest_from_file1)):
        c2(row=1, column=starting_column_file2 + j).value = new_column_name + '_' + str(j)

    # Check all rows using the key column to match key from file1 with key from file 2
    for i in range(2, 10000000):
        key = c2(row=i, column=key_to_file2).value
        if key == None:
            break
        else:
            if key in file1_data:
                for j in range(0, len(file1_data[key])):
                    c2(row=i, column=starting_column_file2 + j).value = file1_data[key][j]

    """
    If the user is interested to add gene_location
    """

    if add_annotation == "No":
        pass

    if add_annotation == "Gene location":
        wb_location = openpyxl.load_workbook('/home/hp/Desktop/COVID blood samples/locations.xlsx', data_only=True)
        sheet_location = wb_location.active
        c_location = sheet_location.cell

        locations = {}
        for i in range(2,38944):
            locations[c_location(row=i, column=1).value] = c_location(row=i, column=2).value

        c2(row=1, column=starting_column_file2 + len(columns_of_interest_from_file1)).value = "Gene location"
        for i in range(2, 1000000):
            key = c2(row=i, column=key_to_file2).value
            if key == None:
                break
            else:
                if key in locations:
                    c2(row=i, column=starting_column_file2+len(columns_of_interest_from_file1)).value = locations[key]

    wb2.save('/home/hp/Desktop/COVID blood samples/Integrated_file_test.xlsx')
    if add_report == True:
        f = open("/home/hp/Desktop/COVID blood samples/Analysis Report.txt", "w+")
        f.write('Analysis Report will be written here')
        f.close()
